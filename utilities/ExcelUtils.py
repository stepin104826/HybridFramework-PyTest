import openpyxl as op

def getRowCount(file, sheetname):
    workbook = op.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def getColCount(file, sheetName):
    workbook = op.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(file, sheetName, rownum, colnum):
    workbook = op.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,colnum).value

def writeData(file,sheetName, rownum, colnum, data):
    workbook = op.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,colnum).value = data
    workbook.save(file)